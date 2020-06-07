import errno
import os
from datetime import datetime, date
import xlrd
import openpyxl as pyxl

TEST_EXCEL_PATH = 'data/Report Data/2020-05-28.xlsx'

def check_dir():
    filename = "data/Receipt Data/" + str(date.today())  # "//Zerozed-pc/shared/DB/temp/resit.txt"
    try:
        # Create target Directory
        os.mkdir(filename)
        print("Directory ", filename, " Created ")
    except FileExistsError:
        print("Directory ", filename, " already exists")

    test = os.path.exists(os.path.dirname('data/Report Data'))
    print(test)
    dirname = os.path.abspath('data/Report Data')

    # Print the directory name
    print(dirname)

    # TODO use this code to get dir size
    def get_directory_size(directory):
        """Returns the `directory` size in bytes."""
        total = 0
        try:
            # print("[+] Getting the size of", directory)
            for entry in os.scandir(directory):
                if entry.is_file():
                    # if it's a file, use stat() function
                    total += entry.stat().st_size
                elif entry.is_dir():
                    # if it's a directory, recursively call this function
                    total += get_directory_size(entry.path)
        except NotADirectoryError:
            # if `directory` isn't a directory, get the file size then
            return os.path.getsize(directory)
        except PermissionError:
            # if for whatever reason we can't open the folder, return 0
            return 0
        return total

    def get_size_format(b, factor=1024, suffix="B"):
        """
        Scale bytes to its proper byte format
        e.g:
            1253656 => '1.20MB'
            1253656678 => '1.17GB'
        """
        for unit in ["", "K", "M", "G", "T", "P", "E", "Z"]:
            if b < factor:
                return f"{b:.2f}{unit}{suffix}"
            b /= factor
        return f"{b:.2f}Y{suffix}"

    size = get_size_format(get_directory_size(dirname))
    print(size)

def check_xlrd():
    wb = xlrd.open_workbook(TEST_EXCEL_PATH)
    sheet = wb.sheet_by_index(0)

    for row_num in range(sheet.nrows):
        row_value = sheet.row_values(row_num)
        print(row_value)

def check_pyxl():
    '''EXCEL FOR REPORT DATA A.Code B.Name C.Qty D.Total E.Avg'''
    theFile = pyxl.load_workbook(TEST_EXCEL_PATH)
    allSheetNames = theFile.sheetnames
    qty_adj = 'C'
    total_adj = 'D'
    avg_adj = 'E'

    for sheet in allSheetNames:
        print("Current sheet name is {}".format(sheet))
        currentSheet = theFile[sheet]

    print("All sheet names {} ".format(theFile.sheetnames))

    def find_specific_cell(value="", show="row"):
        for row in range(1, currentSheet.max_row + 1):
            for column in "ABCDE":  # Here you can add or reduce the columns
                cell_name = "{}{}".format(column, row)

                if currentSheet[cell_name].value == value:
                    # print("{1} cell is located on {0}" .format(cell_name, currentSheet[cell_name].value))
                    # print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))   #debug
                    # print(column, row)
                    if show.lower() == "row":
                        return row
                    elif show.lower() =="column":
                        return column
                    elif show.lower() == "cell":
                        return cell_name
                    else:
                        return 'Only row, column, cell'

    def add_cell(value=['x','x',1,1.1,1.1], column="ABCDE"):
        '''value Must depends on columns like 5 value for "ABCDE" '''

        wb = xlrd.open_workbook(TEST_EXCEL_PATH)
        sheet = wb.sheet_by_index(0)

        check_row = sheet.nrows
        i = 0
        for column in "ABCDE":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, (check_row+1))
            currentSheet[cell_name].value = value[i]
            i+=1

        print('Successfully add new row')
        theFile.save(TEST_EXCEL_PATH)
        theFile.close()


    def get_column_letter(specificCellLetter):
        letter = specificCellLetter[0:-1]
        print(letter)
        return letter

    def get_all_values_by_cell_letter(letter):
        for row in range(1, currentSheet.max_row + 1):
            for column in letter:
                cell_name = "{}{}".format(column, row)
                # print(cell_name)
                print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))


    # print(find_specific_cell(value="Puteri Ayu", show='column'))
    row_adj = str(find_specific_cell(value="Putewri Ayu", show='row'))

    # print(currentSheet[avg_adj+row_adj].value)
    # currentSheet[avg_adj + row_adj].value = currentSheet[avg_adj+row_adj].value + 3

    theFile.save(TEST_EXCEL_PATH)
    theFile.close()

    if row_adj == 'None':
        add_cell(['RZ0000P005','POPIA MUSHROM', 12, 5, 10])


# check_dir()

# check_xlrd()

# check_pyxl()

test_list = [0,2,3,4,5]
print(test_list)
test_list[0] = 'end'
print(test_list)

# for i in range(len(rawValue)):
#     if i >= 1:  # not include CODE1
#         if len(self.list_count_code2) != len(rawValue) - 1:
#             self.list_count_code2.append(0)
#     self.count_code2 = i - 1
#
#     checkDeals(id=rawValue[i][0])
#     print(self.count_code2)
#
#     if self.count_code2 == len(self.list_count_code2):
#         self.count_code2 = 0
