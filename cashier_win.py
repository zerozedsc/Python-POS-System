import errno
import os
import subprocess
import shutil
import sqlite3
import time
import tkinter as tk
import tkinter.simpledialog as sd
from datetime import datetime, date
from sys import exit
from tkinter import *
from tkinter import messagebox
from tkinter.ttk import *
from tkinter.ttk import Style

# import pyttsx3 as tts
import openpyxl as pyxl
import tkcalendar
import win32print
import xlrd
from openpyxl.styles import Alignment
from win32api import GetSystemMetrics

from server import COUNTMAX, LENMAX
# server
from server import SERVER_PATH as PATH


class CashierWin():
    def __init__(self, master, ID=''):
        self.cashWin = master
        self.cashWin.resizable(0, 0)
        self.cashWin.protocol("WM_DELETE_WINDOW", False)
        self.windowHeight = int(GetSystemMetrics(1))  # y
        if 800 < int(GetSystemMetrics(1)):
            self.windowHeight = 800
        self.windowWidth = int(GetSystemMetrics(0))  # x
        if 1400 < int(GetSystemMetrics(0)):
            self.windowWidth = 1400
        self.positionRight = int(self.cashWin.winfo_screenwidth() / 2 - (self.windowWidth / 2))
        self.positionDown = int(self.cashWin.winfo_screenheight() / 2 - (self.windowHeight / 2))
        self.cashWin.iconphoto(False, PhotoImage(file='images/rozeriya.png'))
        self.cashWin.geometry(
            f"{self.windowWidth}x{self.windowHeight}+{self.positionRight}+{self.positionDown}")  # 1400x800
        self.cashWin.title("CASHIER WINDOW")
        style = Style()
        style.configure('W.TButton', font=('Comic sans ms', 15, 'normal', 'italic'), foreground='black')
        style1 = 'W.TButton'

        # Setup Var
        self.IDcashier = ID.upper()
        self.dateNow = datetime.now().strftime("%d/%b/%Y")
        self.dayN = self.k = self.p = self.d = 0
        self.dbPath = PATH
        self.totalMoney = 0
        self.totalBuy = 0
        self.printIntro = True
        self.buy = []
        self.counting = True
        self.cache_code2 = []
        self.cache_code1 = []
        self.memberDeal = []
        self.calcChoice = True
        self.dealDiscount = 0
        self.memberDiscount = 0
        self.showDiscount = 0
        self.calcChange = 0
        self.ID_FOUND = True

        # TTS
        # try:
        #     self.engine = tts.init()
        #     voices = self.engine.getProperty('voices')
        #     self.engine.setProperty('voice', voices[1].id)
        #     rate = self.engine.getProperty('rate')
        #     self.engine.setProperty('rate', 150)
        # except Exception as e:
        #     print(e)

        # Cashier Name
        def name():
            try:
                conn = sqlite3.connect(PATH)
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM EMPLOYEE_DATA ORDER BY EMPLOYEE_ID;")
                query1 = cursor.fetchall()
                conn.commit()
                name = []
                for i in query1:
                    for x in i:
                        name.append(str(x))

                cashierName = 'NAME: ' + str(name[name.index(self.IDcashier.upper()) + 1])
                # nameLen = len(list(map(len, cashierName)))
                # wordLen = list(map(len, cashierName.split(" ")))

                namelbl = Label(self.cashWin, text=cashierName, font=('comic sans ms', 13, 'bold'),
                                foreground='blue')
                namelbl.place(relx=0, rely=0.041)
            except (Exception, sqlite3.Error) as error:
                messagebox.showerror('ERROR', error + ' Please contact Service Team')

        # background
        background_image = tk.PhotoImage(master=self.cashWin, file='images/cashierbg.png')
        background_label = Label(self.cashWin, background='gold', image=background_image)
        background_label.image = background_image
        background_label.place(x=0, y=0, relwidth=1, relheight=1)

        # ID
        self.idCashlbl = Label(self.cashWin, text='ID:ABCX' + self.IDcashier, font=('comic sans ms', 15, 'bold'),
                               foreground='green')
        self.idCashlbl.place(relx=0, rely=0, width=238)
        name()

        # TIME
        self.timeString = StringVar()
        Label(self.cashWin, text='TIME: ', font=('comic sans ms', 12, 'bold'),
              foreground='black').place(relx=0.6, rely=0.041)
        self.timelbl = Label(self.cashWin, text='', textvariable=self.timeString, font=('comic sans ms', 12, 'bold'),
                             foreground='green')
        self.timelbl.place(relx=0.64, rely=0.041, width=238)
        self.timelbl.after(1, self.updateTime)

        # DATE
        Label(self.cashWin, text='DATE: ' + self.dateNow, font=('comic sans ms', 13, 'bold'),
              foreground='black').place(relx=0.6, rely=0.005)

        # back button
        self.backButton = Button(self.cashWin, text='BACK', style=style1, command=self.backChoice)
        self.backButton.place(relx=0.84, rely=0, relwidth=0.08)

        # exit button
        self.exitButton = Button(self.cashWin, text='EXIT', style=style1, command=self.exitWindow)
        self.exitButton.place(relx=0.92, rely=0, relwidth=0.08)

        # total button
        self.totalButton = Button(self.cashWin, text='ENTER TOTAL', style=style1, command=self.payWindow)
        self.totalButton.place(relx=0.48, rely=0.94)

        # cancel sale button
        self.cancelSaleButton = Button(self.cashWin, text='CANCEL SALE', style=style1, command=self.saleCancel)
        self.cancelSaleButton.place(relx=0.48, rely=0.85)

        # print report button
        self.printReportButton = Button(self.cashWin, text='REPORT', style=style1, command=self.printExcel,
                                        state='disabled')
        self.printReportButton.place(relx=0.48, rely=0.65)

        # add new item button
        self.add_item = Button(self.cashWin, text='NEW ITEM', style=style1, command=self.newItemWindow)
        self.add_item.place(relx=0.48, rely=0.45)

        # stock out not in sale button
        self.stock_out = Button(self.cashWin, text='STOCK OUT', style=style1, command=self.stockOut)
        self.stock_out.place(relx=0.48, rely=0.55)

        # Intro
        self.scrollbar = Scrollbar(self.cashWin)
        # self.scrollbar.place(relx=0.8, rely=0.5, height=400)
        self.outputArea = Text(self.cashWin)
        self.outputArea.config(state='disabled', yscrollcommand=self.scrollbar.set)
        self.outputArea.place(relx=0.6, rely=0.2, relwidth=0.4, relheight=0.15)

        # Treeview buy
        cols = ('ID', 'PRODUK', 'QTY', 'HARGA', 'TOTAL', 'OFF')
        self.buyScreen = Treeview(self.cashWin, columns=cols, show='headings')
        i = 0
        for col in cols:
            i = i + 1
            if i == 1:
                self.buyScreen.heading(col, text=col, )
                self.buyScreen.column(col, minwidth=0, width=100, stretch=False)
            elif i == 2:
                self.buyScreen.heading(col, text=col, )
                self.buyScreen.column(col, minwidth=0, width=210, stretch=False)
            elif i == 3:
                self.buyScreen.heading(col, text=col, )
                self.buyScreen.column(col, minwidth=0, width=30, stretch=False)
            elif i == 4:
                self.buyScreen.heading(col, text=col, )
                self.buyScreen.column(col, minwidth=0, width=55, stretch=False)
            elif i == 5:
                self.buyScreen.heading(col, text=col, )
                self.buyScreen.column(col, minwidth=0, width=55, stretch=False)
            elif i == 6:
                self.buyScreen.heading(col, text=col, )
                self.buyScreen.column(col, minwidth=0, width=55, stretch=False)

        # insert to buy

        self.buyScreen.place(relx=0.6, rely=0.35, relwidth=0.4, relheight=0.35)
        vsb = Scrollbar(self.cashWin, orient="vertical", command=self.buyScreen.yview)
        # vsb.place(relx=0.98, rely=0.1, relheight=0.5)
        hrz = Scrollbar(self.cashWin, orient="horizontal", command=self.buyScreen.xview)
        # hrz.place(relx=0.6, rely=0.7, relwidth=0.4)
        self.buyScreen.configure(yscrollcommand=vsb.set, xscrollcommand=hrz.set)

        # TOTAL PRICE
        self.scrollbarPrice = Scrollbar(self.cashWin)
        # self.scrollbar.place(relx=0.8, rely=0.5, height=400)
        self.totalPrice = Text(self.cashWin)
        self.totalPrice.config(state='disabled', yscrollcommand=self.scrollbarPrice.set)
        self.totalPrice.place(relx=0.6, rely=0.7, relwidth=0.4, relheight=0.35)

        # PRODUCT ID
        self.getPrdID = StringVar()
        Label(self.cashWin, text="ID", font=('arial', 15, 'bold')).place(relx=0, rely=0.28)
        self.ID_entry = Entry(self.cashWin, textvariable=self.getPrdID, font=('comic sans ms', 12, 'bold'), state='normal')
        self.ID_entry.place(relx=0, rely=0.32, height=30, width=200)
        self.ID_entry.focus_force()

        # QUANTITY
        self.getPrdQty = IntVar()
        Label(self.cashWin, text="QTY", font=('arial', 15, 'bold')).place(relx=0.18, rely=0.28)
        self.Qty_entry = Entry(self.cashWin, textvariable=self.getPrdQty, font=('comic sans ms', 12, 'bold'),
                               state='disabled')
        self.Qty_entry.place(relx=0.18, rely=0.32, height=30, width=50)

        def q(*args):
            self.Qty_entry.selection_range(0, END)

        self.Qty_entry.bind("<FocusIn>", q)

        # PRICE
        self.getPrdPrice = DoubleVar()
        Label(self.cashWin, text="PRICE", font=('arial', 15, 'bold')).place(relx=0.25, rely=0.28)
        self.Price_entry = Entry(self.cashWin, textvariable=self.getPrdPrice, font=('comic sans ms', 12, 'bold'),
                                 state='disabled')
        self.Price_entry.place(relx=0.25, rely=0.32, height=30, width=100)

        def a(*args):
            self.Price_entry.selection_range(0, END)

        self.Price_entry.bind("<FocusIn>", a)
        self.ID_entry.bind("<Return>", self.productDetail)
        self.cashWin.bind("<Return>", self.insertBuy)

        # cancel button
        self.cancelButton = Button(self.cashWin, text='cancel', command=self.cancel, state='disabled')
        self.cancelButton.place(relx=0.35, rely=0.32)

        try:
            with open('Config/cashierWindowCache.ini', 'r') as cache:
                check_cache = [i.strip("\n") for i in cache.readlines()]
                if str(self.dateNow) == check_cache[0]:
                    if check_cache[1] != self.totalMoney and check_cache[2] != self.dayN:
                        self.totalMoney = float(check_cache[1])
                        self.dayN = int(check_cache[2])

                self.z = int(check_cache[3])
                if self.z >= COUNTMAX:
                    messagebox.askokcancel("Maintenance Required",
                                           "You need a maintenance, Pls Call Service Team at +60172208214")
                    self.cashWin.destroy()
                cache.close()
        except:
            pass

        # buy count and total money
        Label(self.cashWin, text="TOTAL TODAY: RM" + str(self.totalMoney), font=('comic sans ms', 15, 'bold')).place(
            relx=0.6, rely=0.15)
        Label(self.cashWin, text="SALE NO: " + str(self.dayN), font=('comic sans ms', 15, 'bold')).place(relx=0.85,
                                                                                                         rely=0.15)

        # MEMBERCARD ENTRY AND DEAL ENTRY
        self.getMemberID = StringVar()
        Label(self.cashWin, text="MEMBERCARD ID", font=('arial', 15, 'bold'), state='disabled').place(relx=0, rely=0.78)
        self.member_entry = Entry(self.cashWin, textvariable=self.getMemberID, font=('comic sans ms', 12, 'bold'),
                                  state='normal')
        self.member_entry.place(relx=0, rely=0.82, height=30, width=180)

        # def c(*args):
        #     if self.getMemberID.get() != '' and len(self.memberDeal) != 1:
        #         self.memberDeal.append(self.getMemberID.get())
        #         self.member_entry.delete(0, END)
        #         messagebox.showinfo("SUCESS", "MEMBER ID ENTERED")
        #     else:
        #         messagebox.showwarning("EMPTY", "EMPTY ENTRY OR MEMBER ID EXIST")

        def fixed(*args):
            self.ID_entry.focus_force()

        #CheckBox
        self.chkbox1 = IntVar()
        self.chkbox1.set(1)
        self.drawer_box = Checkbutton(self.cashWin, text="Cash Drawer Open",variable=self.chkbox1)
        self.drawer_box.place(relx=0.5, rely=0.2)

        # bind
        self.cashWin.bind("<Up>", self.payWindow)
        self.cashWin.bind("<Down>", self.openDrawer)
        self.cashWin.bind("<Delete>", self.cancel)
        self.cashWin.bind("<End>", self.saleCancel)
        self.cashWin.bind("<Button-1>", fixed)

        self.cashWin.mainloop()

    # receipt
    def receiptIntro(self):
        dirname = "data/Receipt Data/" + str(date.today())  # "//Zerozed-pc/shared/DB/temp/resit.txt"
        try:
            # Create target Directory
            os.mkdir(dirname)
            print("Directory ", dirname, " Created ")  # debug check directory created
        except FileExistsError:
            # print("Directory ", dirname, " already exists")     #debug check dir exists
            pass

        self.outputArea.config(state='normal', yscrollcommand=self.scrollbar.set)
        conn = sqlite3.connect(PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM EMPLOYEE_DATA ORDER BY EMPLOYEE_ID;")
        query1 = cursor.fetchall()
        conn.commit()
        name = []
        for i in query1:
            for x in i:
                name.append(str(x))

        if self.z >= COUNTMAX:
            self.cashWin.destroy()
        elif self.z >= 1000:
            grab_z = str(self.z)
            self.sellID = f"RZ{grab_z[0]}{grab_z[1]}S{grab_z[2]}{grab_z[3]}"
        elif self.z >= 100:
            grab_z = str(self.z)
            self.sellID = f"RZ0{grab_z[0]}S{grab_z[1]}{grab_z[2]}"
        elif self.z >= 10:
            self.sellID = 'RZ00S' + str(self.z)
        else:
            self.sellID = 'RZ00S0' + str(self.z)

        if len(self.sellID) > LENMAX:
            self.cashWin.destroy()
        self.counting = True

        self.filename = dirname + "/" + str(self.sellID) + ".txt"  # "//Zerozed-pc/shared/DB/temp/resit.txt"
        if not os.path.exists(os.path.dirname(self.filename)):
            try:
                os.makedirs(os.path.dirname(self.filename))
            except OSError as exc:  # Guard against race condition
                if exc.errno != errno.EEXIST:
                    raise

        # print(self.n, '', self.sellID)

        def insertData():
            intro = """ROZERIYA ENTERPRISE\nTEL +6017 959 3309""" + \
                    "\n__________________________________________________________" \
                    ""
            resitComp = """__________________________________\n
item\t       Qty   S/Price   Total  Off"""  # 2/slash

            self.outputArea.insert(1.0, intro, 'CENTER')
            date = f"DATE&TIME: {str(self.dateNow)} {str(time.strftime('%H:%M:%S%p'))} \n"
            self.outputArea.insert(INSERT, "")
            self.outputArea.insert(INSERT, f"\nID:{self.IDcashier}\n")
            self.outputArea.insert(INSERT, f"CASHIER:{str(name[name.index(self.IDcashier) + 1])}\n")
            self.outputArea.insert(INSERT, date)
            self.outputArea.insert(INSERT, f"SELL ID:{self.sellID}\n")
            self.outputArea.insert(INSERT, resitComp)
            self.outputArea.config(state='disabled', yscrollcommand=self.scrollbar.set)
            self.introGet = self.outputArea.get(1.0, END)
            with open(self.filename, "w") as f:
                f.write(self.introGet)
                f.close()

        # print(self.outputArea.get(1.0, END))
        insertData()
        # print(self.outputArea.get(1.0, END))

    # total SEND
    def payWindow(self, *args):

        def paying():
            try:
                for i in self.buyScreen.get_children():
                    pass
                if i is not None:
                    self.totalCalc()
                    self.payWin = Toplevel(self.cashWin)
                    self.payWin.lift(aboveThis=self.cashWin)
                    self.payWin.resizable(0, 0)
                    self.payWin.protocol("WM_DELETE_WINDOW", False)
                    self.payWin.iconphoto(False, PhotoImage(file='images/rozeriya.png'))
                    self.payWin.geometry(f"800x500")
                    self.payWin.title("ADMIN FORM")
                    style = Style()
                    style.configure('Fun.TButton', font=('arial', 12, 'normal', 'italic'), foreground='black')
                    style1 = 'Fun.TButton'
                    self.totalButton.config(state='disabled')

                    # background
                    # background_image = tk.PhotoImage(master=self.payWin, file='images/cashierbg.png')
                    background_label = Label(self.payWin, background='silver')
                    # background_label.image = background_image
                    background_label.place(x=0, y=0, relwidth=1, relheight=1)

                    # main frame
                    payFrame = Frame(self.payWin)
                    payFrame.place(relheight=1, relwidth=0.8, relx=0.2)
                    payFrame_bg = Label(payFrame, background='dark blue')
                    payFrame_bg.place(x=0, y=0, relwidth=1, relheight=1)

                    def cash():

                        filename = self.filename  # "//Zerozed-pc/shared/DB/temp/resit.txt"
                        if not os.path.exists(os.path.dirname(filename)):
                            try:
                                os.makedirs(os.path.dirname(filename))
                            except OSError as exc:  # Guard against race condition
                                if exc.errno != errno.EEXIST:
                                    raise
                        Label(payFrame, text="TOTAL:RM" + str(self.actualTotal),
                              font=('comic sans ms', 18, 'bold')).place(relx=0.35, rely=0.15)
                        self.getCash = DoubleVar()
                        Label(payFrame, text="RM:", font=('arial', 20, 'bold')).place(relx=0.15, rely=0.35)
                        self.payCash_entry = Entry(payFrame, textvariable=self.getCash,
                                                   font=('comic sans ms', 20, 'bold'),
                                                   state='enabled')
                        self.payCash_entry.place(relx=0.25, rely=0.35, height=30, width=200)
                        self.payCash_entry.focus_set()

                        def a(*args):
                            if self.getCash.get() >= self.actualTotal:
                                self.calcChange = round(self.getCash.get(), 2) - self.actualTotal
                                self.totalPrice.config(state='normal', yscrollcommand=self.scrollbarPrice.set)
                                self.totalPrice.insert(INSERT, f"CASH \t=\tRM{self.getCash.get()}\n")
                                self.totalPrice.insert(INSERT, f"CHANGE\t=\tRM{round(self.calcChange, 2)}\n")
                                self.totalPrice.insert(INSERT, f"\nTERIMA KASIH • SILA DATANG LAGI ")
                                self.totalGet = self.totalPrice.get(1.0, END)
                                self.totalPrice.config(state='disabled', yscrollcommand=self.scrollbarPrice.set)
                                with open(filename, "a") as f:
                                    f.write(self.totalGet)
                                    f.close()
                                self.totalCmd()
                                self.totalButton.config(state='normal')
                                self.ID_entry.config(state='normal')
                                self.payWin.destroy()
                                self.saveExcel()  # save at excel
                                self.cache_code2.clear()
                                self.cache_code1.clear()
                                self.d = 0
                                self.fix_count = 0
                                self.dayN += 1
                                self.totalMoney = self.totalMoney + self.actualTotal
                                Label(self.cashWin, text="TOTAL TODAY: RM" + str(self.totalMoney),
                                      font=('comic sans ms', 15, 'bold')).place(
                                    relx=0.6, rely=0.15)
                                Label(self.cashWin, text="SALE NO: " + str(self.dayN),
                                      font=('comic sans ms', 15, 'bold')).place(relx=0.85,
                                                                                rely=0.15)

                                with open('Config/cashierWindowCache.ini', 'w+') as cache:
                                    self.z += 1
                                    cache.write(
                                        f"""{self.dateNow}\n{self.totalMoney}\n{self.dayN}\n{self.z}""")  # date,totalMoney, n
                                    cache.close()

                            else:
                                messagebox.showwarning("less cash", "cash is not enough")
                                self.payWin.lift(aboveThis=self.cashWin)

                        self.payCash_entry.bind("<Return>", a)

                        def b(*args):
                            self.payCash_entry.selection_range(0, END)

                        self.payCash_entry.bind("<FocusIn>", b)

                    def back():
                        self.totalButton.config(state='normal')
                        self.ID_entry.config(state='normal')
                        self.totalPrice.config(state='normal', yscrollcommand=self.scrollbarPrice.set)
                        self.totalPrice.delete(1.0, END)
                        self.totalPrice.config(state='disabled', yscrollcommand=self.scrollbarPrice.set)
                        self.payWin.destroy()

                    # CASH BUTTON
                    self.payCash_Button = Button(self.payWin, text='CASH', command=cash, state='enabled', style=style1)
                    self.payCash_Button.place(relx=0, rely=0.15, height=50, width=160)
                    self.payCash_Button.focus_force()

                    # CARD BUTTON
                    self.payCard_Button = Button(self.payWin, text='''CREDIT/DEBIT\nCARD''', command=None,
                                                 state='disabled', style=style1)
                    self.payCard_Button.place(relx=0, rely=0.3, height=50, width=160)

                    # CHEQUE BUTTON
                    self.payCheque_Button = Button(self.payWin, text='CHEQUE', command=None, state='disabled',
                                                   style=style1)
                    self.payCheque_Button.place(relx=0, rely=0.45, height=50, width=160)

                    # CHEQUE BUTTON
                    self.payInstallment_Button = Button(self.payWin, text='INSTALLMENT', command=None, state='disabled',
                                                        style=style1)
                    self.payInstallment_Button.place(relx=0, rely=0.6, height=50, width=160)

                    # BACK BUTTOn
                    self.backPayWin_Button = Button(self.payWin, text='BACK', command=back, state='normal',
                                                    style=style1)
                    self.backPayWin_Button.place(relx=0, rely=0.75, height=50, width=160)

            except Exception as e:
                messagebox.showwarning("PAYING", f"CANNOT PAY\n{e}")

        try:
            for i in self.buyScreen.get_children():
                pass
            if i is not None:
                # if len(self.memberDeal) == 0 or len(self.memberDeal) == 1:
                #     if len(self.memberDeal) == 0:
                #         findMember = messagebox.askquestion("MEMBER ID", "MEMBER ID?", icon='info')
                #         if findMember == 'yes':
                #             findMemberID = sd.askstring(title="MEMBER ID ", prompt="INSERT MEMBER ID")
                #             self.memberDeal.append(str(findMemberID))

                self.fixBuyScreen()
                paying()
                self.cancel()

            else:

                messagebox.showwarning("SALE INPUT", f"NO SALE INPUT TO TOTAL {Exception}")
        except Exception as e:
            messagebox.showwarning("SALE INPUT", f"No Sale or Problem Occur\n{e}")

    def openDrawer(self, *args):
        if self.chkbox1.get() == 1:
            try:
                filename = 'Config/drawer_open.bat'
                dirpath = os.path.abspath(filename)
                subprocess.call([dirpath])
            except (Exception, OSError) as e:
                messagebox.showerror("Cant Open Drawer", f"Drawer Not Found or Any Error\n{e}")

    def totalCmd(self):
        def printResult():
            result = messagebox.askquestion("RECEIPT", f"Baki ialah RM{self.calcChange}, Print Resit?", icon='info')
            if result == 'yes':
                self.openDrawer()
                # print("sent to the printer")  #debug
                self.outputArea.config(state='normal')
                self.outputArea.delete(1.0, END)
                self.outputArea.config(state='disabled', yscrollcommand=self.scrollbar.set)
                self.totalPrice.config(state='normal', yscrollcommand=self.scrollbarPrice.set)
                self.totalPrice.delete(1.0, END)
                self.totalPrice.config(state='disabled', yscrollcommand=self.scrollbarPrice.set)
                for i in self.buyScreen.get_children():
                    self.buyScreen.delete(i)
                printfile = os.path.abspath(self.filename)
                try:
                    try:
                        with open('Config/printerConfig.ini', 'r') as f:
                            check_print = [i.strip("\n") for i in f.readlines()]
                            win32print.SetDefaultPrinter(str(check_print[1]))
                            # print(win32print.GetDefaultPrinter())
                            f.close()
                    except Exception as e:
                        messagebox.showerror("Printer Error",
                                             f"Cant Print, Please contact service provider +60179593309\n{e}")

                    os.startfile(rf"{printfile}", 'print')
                except Exception as error:
                    messagebox.showerror("Printer error", error)
                # try:
                #     self.engine.say("Thank you for buying, and please come again")
                #     self.engine.runAndWait()
                # except Exception as e:
                #     print(e)
            else:
                # print("not sent to the printer")    #debug not sent to the Printer
                self.openDrawer()
                self.outputArea.config(state='normal')
                self.outputArea.delete(1.0, END)
                self.outputArea.config(state='disabled', yscrollcommand=self.scrollbar.set)
                self.totalPrice.config(state='normal', yscrollcommand=self.scrollbarPrice.set)
                self.totalPrice.delete(1.0, END)
                for i in self.buyScreen.get_children():
                    self.buyScreen.delete(i)
                self.totalPrice.config(state='disabled', yscrollcommand=self.scrollbarPrice.set)
                # try:
                #     self.engine.say("Thank you for buying, and please come again")
                #     self.engine.runAndWait()
                # except Exception as e:
                #     print(e)

        self.ID_entry.focus_force()
        self.updateDb()
        self.printIntro = True
        self.counting = True
        Label(self.cashWin, text="SALE NO: " + str(self.dayN), font=('comic sans ms', 15, 'bold')).place(relx=0.85,
                                                                                                         rely=0.15)
        self.buy.clear()
        self.memberDeal.clear()
        self.outputArea.after(1000, printResult)

    def productDetail(self, *args):
        conn = sqlite3.connect(PATH)
        cursor = conn.cursor()
        try:
            cursor.execute(f"SELECT * FROM PRODUCT_DATA WHERE PRODUCT_ID={self.getPrdID.get()};")
            conn.commit()
            query1 = cursor.fetchall()
            product = []
            for i in query1:
                for x in i:
                    product.append(str(x))
            # print(product)

            try:
                self.Qty_entry.config(state='normal')
                self.Price_entry.config(state='normal')
                self.ID_entry.config(state='disabled')
                self.cancelButton.config(state='normal')
                self.Qty_entry.focus_set()
                self.productName = str(product[product.index(self.getPrdID.get()) + 1])
                self.namePrd = Label(self.cashWin,
                                     text="Product Name: " + self.productName + "                       ",
                                     font=('comic sans ms', 15, 'bold'))
                self.namePrd.place(relx=0, rely=0.4)
                productQty = int(product[product.index(self.getPrdID.get()) + 4])
                self.qtyPrd = Label(self.cashWin,
                                    text="KUANTITI: " + str(productQty) + "     ",
                                    font=('comic sans ms', 15, 'bold'))
                if productQty <= 20:
                    self.qtyPrd.config(foreground='red')
                    messagebox.showwarning("Stok", "ISI STOK DENGAN SEGERA")
                elif productQty <= 0:

                    self.qtyPrd.config(foreground='red')
                    messagebox.showwarning("Stok", "ITEM SUDAH KEHABISAN STOK")
                self.qtyPrd.place(relx=0, rely=0.45)
                productPrice = float(product[product.index(self.getPrdID.get()) + 3])
                self.getPrdPrice.set(productPrice)
                self.ID_FOUND = True
                if len(self.outputArea.get("1.0", END)) == 1:
                    self.printIntro = True

            except Exception as e:
                self.getPrdID.set('')
                self.getPrdQty.set(0)
                self.Qty_entry.config(state='disabled')
                self.Price_entry.config(state='disabled')
                self.ID_entry.config(state='normal')
                self.cancelButton.config(state='disabled')
                if self.printIntro:
                    self.printIntro = False
                self.ID_FOUND = False
                messagebox.showerror("ERROR PRODUCT ID 1", f"""PRODUCT ID NOT FOUND OR ENTRY EMPTY\n({e})""")
                self.ID_entry.focus_force()

        except Exception as e:
            self.getPrdID.set('')
            self.getPrdQty.set(0)
            self.Qty_entry.config(state='disabled')
            self.Price_entry.config(state='disabled')
            self.ID_entry.config(state='normal')
            self.cancelButton.config(state='disabled')
            if self.printIntro:
                self.printIntro = False
            self.ID_FOUND = False
            messagebox.showerror("ERROR PRODUCT ID 2", f"""PRODUCT ID NOT FOUND OR ENTRY EMPTY\n({e})""")
            self.ID_entry.focus_force()

    # insert, cancel and calc config
    def insertBuy(self, *args):
        self.showDiscount = 0
        try:
            if self.printIntro:
                self.receiptIntro()
                self.printIntro = False

            if self.ID_FOUND:
                if len(str(self.getPrdQty.get())) >= len(str(self.getPrdID.get())) - 2:
                    # print(">>", self.getPrdID.get())  #DEBUG
                    catch_id = str(self.getPrdQty.get())
                    self.getPrdQty.set(1)
                    priceRound = round(float(self.getPrdPrice.get()), 2)
                    self.calc = self.getPrdQty.get() * round(self.getPrdPrice.get(), 2)
                    self.deals()
                    self.buyScreen.insert("", END,
                                          values=(
                                              self.getPrdID.get(), self.productName, int(self.getPrdQty.get()),
                                              str(priceRound),
                                              str(round(self.calc, 2)), self.showDiscount))
                    self.buy.append(str(self.getPrdID.get()))
                    self.buy.append(str(self.getPrdQty.get()))
                    self.cancel()
                    self.getPrdID.set(str(catch_id))
                    self.getPrdQty.set(1)
                    self.productDetail()

                else:
                    if self.getPrdQty.get() != 0 and self.getPrdID != '':
                        self.ID_entry.focus_force()
                        priceRound = round(float(self.getPrdPrice.get()), 2)
                        self.calc = self.getPrdQty.get() * round(self.getPrdPrice.get(), 2)
                        self.deals()
                        self.buyScreen.insert("", END,
                                              values=(
                                                  self.getPrdID.get(), self.productName, int(self.getPrdQty.get()),
                                                  str(priceRound),
                                                  str(round(self.calc, 2)), self.showDiscount))
                        self.buy.append(str(self.getPrdID.get()))
                        self.buy.append(str(self.getPrdQty.get()))
                        self.cancel()

                    else:
                        self.getPrdQty.set(1)  # debug blank
            else:
                messagebox.showerror("ID NOT FOUND", "PLS SCAN OR ENTER PRODUCT ID AGAIN")


        except Exception as e:
            messagebox.showerror("WRONG INPUT", "TRY AGAIN")
            print(e)

    def totalCalc(self, *args):
        filename = self.filename
        if not os.path.exists(os.path.dirname(filename)):
            try:
                os.makedirs(os.path.dirname(filename))
            except OSError as exc:  # Guard against race condition
                if exc.errno != errno.EEXIST:
                    raise
        component = []
        totalPrice = 0.00
        for child in self.buyScreen.get_children():
            component.append(
                float(self.buyScreen.item(child)["values"][4]) + float(self.buyScreen.item(child)["values"][5]))
        for i in component:
            # print(component)  #debug
            totalPrice += float(i)
        calcRounding = round(totalPrice - round(totalPrice, 1), 2)
        if calcRounding < 0:
            a = ''
        else:
            a = '-'
        self.totalPrice.config(state='normal', yscrollcommand=self.scrollbarPrice.set)
        self.totalPrice.insert(INSERT, f"\nSUBTOTAL(+OFF):\tRM{totalPrice}\n")
        self.totalPrice.insert(INSERT, f"ROUNDING:\tRM{a}{calcRounding}\n")
        self.totalPrice.insert(INSERT, "========================================\n ")
        tax = open("data/tax.cfg")
        n = ''
        for i in tax.read():
            n = n + i
        self.actualTotal = totalPrice - calcRounding - self.dealDiscount
        # try:
        #     self.engine.say(f"Your Total is {self.actualTotal} Ringgit")
        #     self.engine.runAndWait()
        # except Exception as e:
        #     print(e)
        self.totalPrice.insert(INSERT, f"\nTAX: {n} %\n")
        self.totalPrice.insert(INSERT, f"DISCOUNT =\t-RM{self.dealDiscount}\n")
        self.totalPrice.insert(INSERT, f"TOTAL\t=\tRM{round(self.actualTotal, 2)}\n")
        # self.totalPrice.insert(INSERT, f"MEMBER ID:{self.memberCard()}\n")
        self.totalPrice.config(state='disabled', yscrollcommand=self.scrollbarPrice.set)

        self.dealDiscount = 0

    def updateDb(self, *args):
        try:
            n = 0
            p = 1
            for i in range(0, len(self.buy)):
                if len(self.buy) > n:
                    conn = sqlite3.connect(PATH)
                    cursor = conn.cursor()
                    cursor.execute(f"SELECT QUANTITY FROM PRODUCT_DATA WHERE PRODUCT_ID = '{self.buy[n]}'")
                    conn.commit()
                    rawValue = cursor.fetchone()
                    for i in rawValue:
                        count = int(i) - int(self.buy[p])
                    cursor.execute(f"UPDATE PRODUCT_DATA SET QUANTITY=? WHERE PRODUCT_ID = ?", (count, self.buy[n],))
                    conn.commit()
                    p = p + 2
                    n = n + 2
        except (Exception, sqlite3.Error) as e:
            messagebox.showerror("ERROR UPDATE DATABASE" f"ERROR WHEN UPDATE DATABASE\n{e}")

    def saleCancel(self, *args):
        try:
            self.payWin.destroy()
        except:
            pass

        self.ID_entry.focus_force()
        self.counting = False
        self.fix_count = 0
        self.cache_code2.clear()
        self.cache_code1.clear()
        self.showDiscount = 0
        try:
            for i in self.buyScreen.get_children():
                self.buyScreen.delete(i)
            if i is not None:
                self.memberDeal.clear()
                self.outputArea.config(state='normal')
                self.outputArea.delete(1.0, END)
                self.outputArea.config(state='disabled', yscrollcommand=self.scrollbar.set)
                self.totalPrice.config(state='normal', yscrollcommand=self.scrollbarPrice.set)
                self.totalPrice.delete(1.0, END)
                self.totalPrice.config(state='disabled', yscrollcommand=self.scrollbarPrice.set)
                self.printIntro = True
                self.cancel()
            else:
                messagebox.showwarning("SALE INPUT", "NO SALE INPUT TO DELETE")
        except:
            messagebox.showwarning("SALE INPUT", "NO SALE INPUT TO DELETE")

    def newItemWindow(self):
        conn = sqlite3.connect(PATH)  # \\Zerozed-pc\shared\DB
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM PRODUCT_DATA ORDER BY PRODUCT_ID;")
        query = cursor.fetchall()
        conn.commit()
        # pprint(query) #debug

        try:
            check_id = False
            self.reg = sd.askstring('REGISTER ITEM', f'PRODUCT ID ')
            if isinstance(self.reg, int):
                self.reg = str(self.reg).strip()
            else:
                self.reg = str(self.reg).strip().upper()

            if self.reg is None:
                pass
            else:
                for i in query:
                    if self.reg in i:
                        check_id = True
                        break

                if not check_id:
                    if "RZ" and 'P' in self.reg or len(self.reg) >= 10:
                        self.addWin = Toplevel(self.cashWin)
                        self.addWin.resizable(0, 0)
                        self.windowHeight = int(self.addWin.winfo_reqheight())
                        self.windowWidth = int(self.addWin.winfo_reqwidth())
                        self.positionRight = int(self.addWin.winfo_screenwidth() / 2 - (self.windowWidth / 2))
                        self.positionDown = int(self.addWin.winfo_screenheight() / 2 - (self.windowHeight / 2))
                        self.addWin.iconphoto(False, PhotoImage(file='images/rozeriya.png'))
                        self.addWin.geometry(f"400x400+{self.positionRight - 400}+{self.positionDown - 300}")
                        self.addWin.title("ADD ITEM")

                        Label(self.addWin, text="REGISTER ITEM", font=('comic sans ms', 18, 'italic', 'bold')).pack()


                        self.getName = StringVar()
                        self.getType = StringVar()
                        self.getPrice = DoubleVar()
                        self.getQty = IntVar()

                        # ID DISABLED
                        Label(self.addWin, text='PRODUCT ID',
                              font=('Comic sans ms', 12, 'normal', 'italic')).place(
                            relx=0,
                            rely=0.2)
                        reg_entry = Label(self.addWin, font=('Comic sans ms', 12, 'normal', 'italic'),
                                          text=self.reg)
                        reg_entry.place(relx=0.38, rely=0.2, width=240)

                        # NAME
                        Label(self.addWin, text='PRODUCT NAME',
                              font=('Comic sans ms', 12, 'normal', 'italic')).place(relx=0, rely=0.3)
                        name_entry = Entry(self.addWin, textvariable=self.getName)
                        name_entry.place(relx=0.38, rely=0.3, width=240)

                        # JENIS
                        Label(self.addWin, text='TYPE OF PRODUCT',
                              font=('Comic sans ms', 12, 'normal', 'italic')).place(relx=0, rely=0.4)
                        type_entry = Entry(self.addWin, textvariable=self.getType)
                        type_entry.place(relx=0.38, rely=0.4, width=240)

                        # PRICE
                        Label(self.addWin, text='PRICE',
                              font=('Comic sans ms', 12, 'normal', 'italic')).place(relx=0, rely=0.5)
                        type_entry = Entry(self.addWin, textvariable=self.getPrice)
                        type_entry.place(relx=0.38, rely=0.5, width=240)

                        # QUANTITY
                        Label(self.addWin, text='QUANTITY',
                              font=('Comic sans ms', 12, 'normal', 'italic')).place(relx=0, rely=0.6)
                        type_entry = Entry(self.addWin, textvariable=self.getQty)
                        type_entry.place(relx=0.38, rely=0.6, width=240)

                        def query_update():
                            if isinstance(self.reg, int):
                                ID = str(self.reg)
                            else:
                                ID = str(self.reg).upper()

                            JENIS = self.getType.get().upper()
                            NAMA = self.getName.get().upper()
                            HARGA = self.getPrice.get()
                            STOCK = self.getQty.get()
                            d = datetime.now()
                            self.timestamp = d.timestamp()
                            try:
                                conn = sqlite3.connect(PATH)  # \\Zerozed-pc\shared\DB
                                cursor = conn.cursor()
                                # print("YOU ARE CONNECTED TO DATABASE")   #debug
                                # print(ID, NAMA, JENIS, HARGA, STOCK)  #debug
                                try:
                                    cursor.execute("SELECT REGISTER_ID FROM REGISTER_CACHE")
                                    reg_table = cursor.fetchall()
                                    conn.commit()
                                    reg_id_count = 1
                                    for i in reg_table:
                                        reg_id_count = reg_id_count + 1
                                    if reg_id_count < 10:
                                        reg_id = "RZ0000R00" + str(reg_id_count)
                                    elif reg_id_count < 100:
                                        reg_id = "RZ0000R0" + str(reg_id_count)
                                    else:
                                        reg_id = "RZ0000R" + str(reg_id_count)
                                    cursor.execute("INSERT INTO PRODUCT VALUES(?,?,?)", (ID, JENIS, True))
                                    conn.commit()
                                    cursor.execute("INSERT INTO PRODUCT_DATA VALUES(?,?,?,?,?)",
                                                   (ID, NAMA, JENIS, HARGA, STOCK))
                                    conn.commit()
                                    cursor.execute("INSERT INTO REGISTER_CACHE VALUES(?,?,?)",
                                                   (reg_id, ID, self.timestamp))
                                    conn.commit()
                                    messagebox.showinfo("SUCCESS", f"REGISTER {ID} DONE")
                                    self.addWin.destroy()
                                except:
                                    messagebox.showerror("FAILED", "ID ALREADY REGISTERED")

                            except (Exception, sqlite3.Error) as error:
                                self.check = False
                                messagebox.showerror("FAILED", f"NO DATABASE DETECTED\n{error}")
                                self.addWin.destroy()


                        # reg button
                        reg_button = Button(self.addWin, text='REGISTER',
                                                    command=query_update)
                        reg_button.place(relx=0.5, rely=0.7, height=100,width=100)


                    else:
                        messagebox.showerror("WRONG ID", "PUT THE CORRECT PRODUCT ID")

                else:
                    messagebox.showerror("WRONG ID", "ID Already Exist")

        except Exception as e:
            messagebox.showwarning("ERROR", f"error: {e}")

    def stockOut(self):
        self.win_stockOut = Toplevel(self.cashWin)
        self.win_stockOut.resizable(0, 0)
        self.win_stockOut.protocol("WM_DELETE_WINDOW", False)
        self.win_stockOut.title("CASHIER WINDOW")
        self.win_stockOut.geometry("600x600")
        self.stock_out.config(state='disabled')
        #bg
        background_image = tk.PhotoImage(master=self.win_stockOut, file='images/cashierbg.png')
        background_label = Label(self.win_stockOut, background='gold', image=background_image)
        background_label.image = background_image
        background_label.place(x=0, y=0, relwidth=1, relheight=1)

        Label(self.win_stockOut, text='Stock Out', font=('comic sans ms', 20, 'bold')).pack()

        #TreeView
        # Treeview buy
        cols = ('ID', 'PRODUK', 'QTY')
        treeview_outStock = Treeview(self.win_stockOut, columns=cols, show='headings')
        i = 0
        for col in cols:
            i += 1
            if i == 1:
                treeview_outStock.heading(col, text=col, )
                treeview_outStock.column(col,minwidth=0, width=150, stretch=False)
            elif i == 2:
                treeview_outStock.heading(col, text=col, )
                treeview_outStock.column(col, minwidth=0, width=350,stretch=False)
            elif i == 3:
                treeview_outStock.heading(col, text=col, )
                treeview_outStock.column(col,minwidth=0, width=100, stretch=False)

        treeview_outStock.place(relx=0, rely=0.1, relwidth=1, relheight=0.4)
        vsb = Scrollbar(self.win_stockOut, orient="vertical", command=treeview_outStock.yview)
        # vsb.place(relx=0.98, rely=0.1, relheight=0.5)
        hrz = Scrollbar(self.win_stockOut, orient="horizontal", command=treeview_outStock.xview)
        # hrz.place(relx=0.6, rely=0.7, relwidth=0.4)
        treeview_outStock.configure(yscrollcommand=vsb.set, xscrollcommand=hrz.set)

        # PRODUCT ID
        getPrdID = StringVar()
        Label(self.win_stockOut, text="ID", font=('arial', 15, 'bold')).place(relx=0, rely=0.52)
        id_entry = Entry(self.win_stockOut, textvariable=getPrdID, font=('comic sans ms', 12, 'bold'),
                              state='normal')
        id_entry.place(relx=0, rely=0.58, height=30, width=200)
        id_entry.focus_force()

        # QUANTITY
        getPrdQty = IntVar()
        getPrdQty.set(1)
        Label(self.win_stockOut, text="QTY", font=('arial', 15, 'bold')).place(relx=0.38, rely=0.52)
        qty_entry = Entry(self.win_stockOut, textvariable=getPrdQty, font=('comic sans ms', 12, 'bold'),
                               state='disabled')
        qty_entry.place(relx=0.38, rely=0.58, height=30, width=50)


        #Function
        def updateDb():
            update_done = False
            stock = []
            for child in treeview_outStock.get_children():
                stock.append([treeview_outStock.item(child)["values"][i] for i in range(3)])

            for g in stock:
                try:
                    conn = sqlite3.connect(PATH)
                    cursor = conn.cursor()
                    cursor.execute(f"SELECT QUANTITY FROM PRODUCT_DATA WHERE PRODUCT_ID = '{g[0]}'")
                    conn.commit()
                    rawValue = cursor.fetchone()
                    print(rawValue, len(rawValue))
                    count = int(rawValue[0]) - int(g[2])
                    cursor.execute(f"UPDATE PRODUCT_DATA SET QUANTITY=? WHERE PRODUCT_ID = ?", (count, g[0],))
                    conn.commit()
                    update_done = True
                except Exception as e:
                    messagebox.showerror("Error", e)

            if update_done:
                for child in treeview_outStock.get_children():
                    treeview_outStock.delete(child)
                messagebox.showinfo("Done Update Stock", 'Stock Successfully Updated')
                id_entry.focus_force()

        def ID_FOUND(*args):
            conn = sqlite3.connect(PATH)
            cursor = conn.cursor()
            cursor.execute(f"SELECT * FROM PRODUCT_DATA")
            conn.commit()
            rawValue = cursor.fetchall()
            found = []

            for i in rawValue:
                if getPrdID.get() in i:
                    found = list(i)
                    break

            if found:
                found.append(True)
                return found
            else:
                found.append(False)
                return found

        def productDetail(*args):
            global namePrd, qtyPrd
            conn = sqlite3.connect(PATH)
            cursor = conn.cursor()
            try:
                cursor.execute(f"SELECT * FROM PRODUCT_DATA WHERE PRODUCT_ID={getPrdID.get()};")
                conn.commit()
                query1 = cursor.fetchall()
                product = []
                for i in query1:
                    for x in i:
                        product.append(str(x))
                # print(product)

                try:
                    qty_entry.config(state='normal')

                    def q(*args):
                        qty_entry.selection_range(0, END)

                    qty_entry.bind("<FocusIn>", q)

                    id_entry.config(state='disabled')
                    cancelButton.config(state='normal')
                    qty_entry.focus_set()
                    productName = str(product[product.index(getPrdID.get()) + 1])
                    namePrd = Label(self.win_stockOut,
                                         text="Product Name: " + productName + "                       ",
                                         font=('comic sans ms', 15, 'bold'))
                    namePrd.place(relx=0, rely=0.65)
                    productQty = int(product[product.index(getPrdID.get()) + 4])
                    qtyPrd = Label(self.win_stockOut,
                                        text="KUANTITI: " + str(productQty) + "     ",
                                        font=('comic sans ms', 15, 'bold'))
                    if productQty <= 20:
                        qtyPrd.config(foreground='red')
                        messagebox.showwarning("Stok", "ISI STOK DENGAN SEGERA")
                    elif productQty <= 0:
                        qtyPrd.config(foreground='red')
                        messagebox.showwarning("Stok", "ITEM SUDAH KEHABISAN STOK")
                    qtyPrd.place(relx=0, rely=0.7)


                except Exception as e:
                    getPrdID.set('')
                    getPrdQty.set(0)
                    qty_entry.config(state='disabled')
                    id_entry.config(state='normal')
                    cancelButton.config(state='disabled')
                    messagebox.showerror("ERROR PRODUCT ID 1", f"""PRODUCT ID NOT FOUND OR ENTRY EMPTY\n({e})""")
                    id_entry.focus_force()

            except Exception as e:
                getPrdID.set('')
                getPrdQty.set(0)
                qty_entry.config(state='disabled')
                id_entry.config(state='normal')
                cancelButton.config(state='disabled')
                messagebox.showerror("ERROR PRODUCT ID 2", f"""PRODUCT ID NOT FOUND OR ENTRY EMPTY\n({e})""")
                id_entry.focus_force()

        def insertStock(*args):

            try:
                catch = ID_FOUND()
                if True in catch:
                    if len(str(getPrdQty.get())) >= len(str(getPrdID.get())) - 2:
                        # print(">>", self.getPrdID.get())  #DEBUG
                        catch_id = str(getPrdQty.get())
                        getPrdQty.set(1)
                        treeview_outStock.insert("", END,values=(getPrdID.get(), catch[1] , int(getPrdQty.get())))
                        cancel()
                        getPrdID.set(str(catch_id))
                        getPrdQty.set(1)
                        productDetail()

                    else:
                        if getPrdQty.get() != 0 and getPrdID != '':
                            treeview_outStock.insert("", END, values=(getPrdID.get(), catch[1] , int(getPrdQty.get())))
                            cancel()

                        else:
                            getPrdQty.set(1)  # debug blank
                else:
                    messagebox.showerror("ID NOT FOUND", "PLS SCAN OR ENTER PRODUCT ID AGAIN")


            except Exception as e:
                messagebox.showerror("WRONG INPUT", "TRY AGAIN")
                print(e)

        def cancel():
            try:
                qty_entry.config(state='disabled')
                id_entry.config(state='normal')
                id_entry.focus_force()
                namePrd.destroy()
                qtyPrd.destroy()
                cancelButton.config(state='disabled')
                getPrdID.set('')
                getPrdQty.set(0)

            except Exception as e:
                messagebox.showerror("Error", f"No Product ID Input In ID Entry\n{e}")

        def quit():
            if len(treeview_outStock.get_children()) != 0:
                ui = messagebox.askyesno("Urusan Belum Selesai", "Anda Pasti Mahu Keluar?")
                if ui:
                    self.stock_out.config(state='normal')
                    self.win_stockOut.destroy()
                else:
                    id_entry.focus_force()
            else:
                self.stock_out.config(state='normal')
                self.win_stockOut.destroy()


        # exit button
        exitButton = Button(self.win_stockOut, text='EXIT', style='W.TButton',
                            command=quit)
        exitButton.place(relx=0.84, rely=0.94, width=100)

        # cancel button
        cancelButton = Button(self.win_stockOut, text='cancel', command=None, state='disabled')
        cancelButton.place(relx=0.62, rely=0.58)

        # total button
        totalButton = Button(self.win_stockOut, text='ENTER TOTAL', style='W.TButton', command=updateDb)
        totalButton.place(relx=0, rely=0.94)

        #bind
        id_entry.bind("<Return>", productDetail)
        qty_entry.bind("<Return>", insertStock)

    # data manipulate config
    def memberCard(self, *args, id=''):
        if id == '':
            pass
        else:
            try:
                try:
                    conn = sqlite3.connect(PATH)
                    cursor = conn.cursor()
                    cursor.execute(f"SELECT * FROM MEMBER WHERE MEMBER_ID = '{id}'")
                    conn.commit()
                    rawValue = cursor.fetchone()

                    # print(rawValue)   #debug Search rawValue

                    # 3 types of members Premium, Gold, Bronze
                    # premium discount -
                    # gold discount -
                    # bronze discount -

                    def Premium():
                        print("premium member")  # debug

                        rawItem = []
                        getType = []  # [0] id, [1] quantity,[2] price ,[3] total
                        n = 0
                        p = 0
                        for child in self.buyScreen.get_children():
                            for i in self.buyScreen.item(child)["values"]:
                                rawItem.append(i)
                        for a in range(0, len(rawItem)):
                            if len(rawItem) > n:
                                cursor.execute(f"SELECT * FROM PRODUCT WHERE PRODUCT_ID = '{rawItem[n]}'")
                                conn.commit()

                    try:
                        if rawValue[1] == "PREMIUM":
                            Premium()
                    except:
                        messagebox.showerror("MEMBER ID", "MEMBER ID NOT FOUND")


                except Exception as e:
                    messagebox.showerror("ERROR", f"MEMBER ID NOT FOUND")
            except Exception as e:
                messagebox.showerror("ERROR", f"ERROR: {e}")

    def deals(self, *args):  # remove old def deals
        conn = sqlite3.connect(PATH)
        cursor = conn.cursor()

        def checkDeals(id='', type=''):
            cursor.execute(f"SELECT * FROM DEAL_DATA WHERE DEAL_ID = '{id}'")
            conn.commit()
            rawValue = cursor.fetchone()

            # print(rawValue)   #debug

            # @ BELI X DAPAT Y
            def CODE1():
                try:
                    if 'BELI' in str(rawValue[6]) and 'MANA' in str(rawValue[6]) and 'DAPAT' in str(rawValue[6]):
                        # print("CODE 1")   #debug Check Code
                        rawItem = [self.getPrdID.get(), self.productName, str(self.getPrdQty.get()),
                                   str(round(float(self.getPrdPrice.get()), 2)),
                                   str(self.getPrdQty.get() * round(self.getPrdPrice.get(), 2))]
                        getType = []  # [0] id, [1] quantity,[2] price ,[3] total
                        n = 0
                        p = 0

                        for a in range(0, len(rawItem)):
                            if len(rawItem) > n:
                                cursor.execute(f"SELECT * FROM PRODUCT WHERE PRODUCT_ID = '{rawItem[n]}'")
                                conn.commit()
                                # rawValue[1] = product_types, rawValue[2] = product_name
                                if rawValue[2] != None:
                                    if cursor.fetchone()[1] == str(rawValue[2]):
                                        getType.append(rawItem[n])
                                        getType.append(rawItem[n + 2])

                                elif rawValue[1] != None:
                                    if cursor.fetchone()[1] == str(rawValue[1]):
                                        # print(db.cursor.fetchall())   #check db fetchall
                                        getType.append(rawItem[n])  # prd id
                                        getType.append(rawItem[n + 2])  # prd qty
                                        getType.append(rawItem[n + 3])  # prd price
                                        getType.append(rawItem[n + 4])  # prd total

                                n = n + 5

                        n = 0
                        for i in range(0, len(getType)):
                            if len(getType) > n:
                                # print('buy 3 :', getType[n + 1])
                                p = p + int(getType[n + 1])
                                n = n + 4

                        x = int(rawValue[3])
                        y = float(rawValue[4])
                        toDiscount = (p - (p % x))

                        # if p < x, Save in self.cache_code1, Next Calc until found discount
                        if toDiscount == 0:
                            self.cache_code1.append([self.getPrdID.get(), self.getPrdQty.get()])
                            # print("P<X", self.cache_code1, len(self.cache_code1))
                            code1Discount = 0
                            check_discount = 0
                            rangeA = len(self.cache_code1)
                            if rangeA > 1:
                                for g in range(rangeA):
                                    check_discount += self.cache_code1[g][1]
                                    k = 0
                                    if check_discount > x:
                                        k = check_discount
                                        check_discount = x
                                    if check_discount % x == 0:
                                        code1Discount += ((float(getType[2]) * x) - y)
                                        check_discount = k - x
                                self.cache_code1.clear()
                                if check_discount > 0:
                                    self.cache_code1.append([self.getPrdID.get(), check_discount])
                                    # print("SAVE ", self.cache_code1)
                            self.showDiscount = code1Discount


                        # If toDiscount == p which mean equal to x, normal calculate
                        elif toDiscount == p:
                            # print("0", self.cache_code1)
                            code1Discount = (toDiscount / x) * ((float(getType[2]) * x) - y)
                            self.showDiscount = code1Discount

                        # if p > x, Reminder save into cache
                        else:
                            # print("P>X", self.cache_code1)
                            code1Discount = 0
                            check_discount = 0
                            for c in range(1, p):
                                if c % x == 0:
                                    code1Discount += ((float(getType[2]) * x) - y)
                                    # print(code1Discount)
                            self.cache_code1.append([self.getPrdID.get(), p % x])
                            rangeA = len(self.cache_code1)
                            if rangeA > 1:
                                for i in range(rangeA):
                                    check_discount += self.cache_code1[i][1]
                                    k = 0
                                    if check_discount > x:
                                        k = check_discount
                                        check_discount = x
                                    if check_discount % x == 0:
                                        code1Discount += (toDiscount / x) * ((float(getType[2]) * x) - y)
                                        check_discount = k - x
                                self.cache_code1.clear()
                                if check_discount > 0:
                                    self.cache_code1.append([self.getPrdID.get(), check_discount])

                            self.showDiscount = code1Discount

                        # print(f"n {n}, p {p}, x {x}, y {y}")
                        # print(f"Discount: {code1Discount}, toDiscount: {toDiscount}")
                        # print(code1Discount,'BELI ' + str(x) ,'DAPAT ' + str(y))  #debug Tell About The Code
                    else:
                        # print("find 1 else")   #debug
                        pass
                except Exception as e:
                    # print("find 1 : ", e)  # debug
                    pass

            # @ BELI LEBIH X 1 DAPAT Y
            def CODE2():
                try:
                    if 'LEBIH' in str(rawValue[6]) and 'DAPAT' in str(rawValue[6]):
                        # print("CODE 2")   #debug
                        rawItem = [self.getPrdID.get(), self.productName, str(self.getPrdQty.get()),
                                   str(round(float(self.getPrdPrice.get()), 2)),
                                   str(self.getPrdQty.get() * round(self.getPrdPrice.get(), 2))]
                        getType = []  # [0] id, [1] quantity,[2] price ,[3] total
                        n = 0
                        p = 0

                        for a in range(0, len(rawItem)):
                            if len(rawItem) > n:
                                cursor.execute(f"SELECT * FROM PRODUCT WHERE PRODUCT_ID = '{rawItem[n]}'")
                                conn.commit()
                                # rawValue[1] = product_types, rawValue[2] = product_name
                                if rawValue[2] != None:
                                    if cursor.fetchone()[1] == str(rawValue[2]):
                                        getType.append(rawItem[n])
                                        getType.append(rawItem[n + 2])

                                elif rawValue[1] != None:
                                    if cursor.fetchone()[1] == str(rawValue[1]):
                                        getType.append(rawItem[n])  # prd id
                                        getType.append(rawItem[n + 2])  # prd qty
                                        getType.append(rawItem[n + 3])  # prd price
                                        getType.append(rawItem[n + 4])  # prd total

                                n = n + 5
                        # print(getType)    #debug
                        n = 0
                        for i in range(0, len(getType)):
                            if len(getType) > n:
                                p += int(getType[n + 1])
                                n = n + 4

                        k = 0
                        x = int(rawValue[3])
                        y = float(rawValue[4])

                        for c in self.cache_code2:
                            if type in c:
                                # print('RUN FOR ', type, c, self.cache_code2.index(c))
                                if self.cache_code2[self.cache_code2.index(c)][1] == 'end':
                                    k = k + (float(getType[2]) - y)
                                    toDiscount = k * int(getType[1])
                                    # print(code1Discount, 'LEBIH ' + str(x), '1 DAPAT ' + str(y))  #debug
                                    self.showDiscount = toDiscount
                                    # self.dealDiscount += toDiscount
                                    break
                                else:
                                    self.cache_code2[self.cache_code2.index(c)][1] += int(getType[1])

                                if p >= x or int(self.cache_code2[self.cache_code2.index(c)][1]) >= x:
                                    k = k + (float(getType[2]) - y)
                                    toDiscount = k * int(getType[1])
                                    # print(code1Discount, 'LEBIH ' + str(x), '1 DAPAT ' + str(y))  #debug
                                    self.showDiscount = toDiscount
                                    # self.dealDiscount += toDiscount
                                    if int(self.cache_code2[self.cache_code2.index(c)][1]) >= x:
                                        self.cache_code2[self.cache_code2.index(c)][1] = 'end'
                                    break

                                else:
                                    k = k + (float(getType[2]) - y)
                                    self.cache_code2.append([self.getPrdID.get(), k * int(getType[1])])
                                    break



                    else:
                        # print("find 2 else")    #debug
                        pass
                except Exception as e:
                    # print(n, k, x, y)   #debug Show n, x, k, y value Ignored
                    # print('find 2 :', e)  # debug
                    pass

            CODE1()
            CODE2()

        try:
            cursor.execute(f"SELECT * FROM DEAL")
            conn.commit()
            fetchValue = cursor.fetchall()
            cursor.execute(f"SELECT * FROM PRODUCT_DATA WHERE PRODUCT_ID = '{self.getPrdID.get()}'")
            conn.commit()
            catching_data = cursor.fetchone()
            # print(rawValue)   #debug
            # print(catch_data)

            if self.cache_code2 == []:
                for i in fetchValue:
                    self.cache_code2.append([i[1], 0])

            # print("found", self.cache_code2)

            for i in fetchValue:
                if catching_data[2] == i[1]:
                    checkDeals(i[0], catching_data[2])
                    break

            # print(self.cache_code2)

            # TODO Need to fix more, slack everywhere

        except Exception as e:
            messagebox.showerror("ERROR", f"ERROR: {e}")

    def fixBuyScreen(self):
        conn = sqlite3.connect(PATH)
        cursor = conn.cursor()
        base = []
        calc_disc = []

        for child in self.buyScreen.get_children():
            base.append([self.buyScreen.item(child)["values"][i] for i in range(6)])
            self.buyScreen.delete(child)

        if self.cache_code2 != []:
            # print(self.cache_code2)
            for cache in self.cache_code2:
                if 'end' in cache:
                    for one in base:
                        cursor.execute(f"SELECT * FROM PRODUCT_DATA WHERE PRODUCT_ID = '{one[0]}'")
                        conn.commit()
                        catch_data = cursor.fetchone()
                        if cache[0] == catch_data[2]:
                            for check_cache in self.cache_code2:
                                if check_cache[0] == str(one[0]):
                                    if base[base.index(one)][5] == 0:
                                        base[base.index(one)][5] = float(check_cache[1])
                                        self.cache_code2.pop(self.cache_code2.index(check_cache))

        for x in base:
            base[base.index(x)][4] = float(base[base.index(x)][4]) - float(base[base.index(x)][5])

        for g in base:
            for p in base:
                if base[base.index(g)] == base[base.index(p)]:
                    pass
                elif g[0] in base[base.index(p)]:
                    base[base.index(g)][5] = float(base[base.index(g)][5]) + float(base[base.index(p)][5])
                    base[base.index(g)][4] = float(base[base.index(g)][4]) + float(base[base.index(p)][4])
                    base[base.index(g)][2] = int(base[base.index(g)][2]) + int(base[base.index(p)][2])
                    base.pop(base.index(p))
                    # print(base)   #debug

        for t in base:
            self.buyScreen.insert("", END, value=t)

        for child1 in self.buyScreen.get_children():
            calc_disc.append([self.buyScreen.item(child1)["values"][i] for i in range(6)])

        for k in calc_disc:
            with open(self.filename, "a") as f:
                productApd = f"""{calc_disc[calc_disc.index(k)][1]}\n{calc_disc[calc_disc.index(k)][0]}  {str(calc_disc[calc_disc.index(k)][2])}  RM{calc_disc[calc_disc.index(k)][3]}  RM{calc_disc[calc_disc.index(k)][4]} -{calc_disc[calc_disc.index(k)][5]}\n"""
                f.write(productApd)
                f.close()
            self.dealDiscount += float(k[5])
        # pprint(base)
        # pprint(self.cache_code2)  #debug

    # misc config
    def updateTime(self):
        self.timeGet = time.strftime('%H:%M:%S %p')
        self.timeString.set(self.timeGet)
        self.timelbl.after(1000, self.updateTime)

    def cancel(self, *args):
        try:
            self.Qty_entry.config(state='disabled')
            self.Price_entry.config(state='disabled')
            self.ID_entry.config(state='normal')
            self.ID_entry.focus_force()
            self.namePrd.destroy()
            self.qtyPrd.destroy()
            self.cancelButton.config(state='disabled')
            self.getPrdID.set('')
            self.getPrdPrice.set(0)
            self.getPrdQty.set(0)

        except Exception as e:
            messagebox.showerror("Error", f"No Product ID Input In ID Entry\n{e}")


        # print("destroy")  #debug

    def backChoice(self):
        # print("Back") #debug
        with open('Config/cashierWindowCache.ini', 'w+') as cache:
            cache.write(f"""{self.dateNow}\n{self.totalMoney}\n{self.dayN}\n{self.z}""")  # date,totalMoney, n
            cache.close()

        self.checkExcel()
        self.saveExit()
        from session_start import SessionStart
        self.cashWin.destroy()
        start_choice = SessionStart().auth(Tk())
        return start_choice

    def exitWindow(self):
        with open('Config/cashierWindowCache.ini', 'w+') as cache:
            cache.write(f"""{self.dateNow}\n{self.totalMoney}\n{self.dayN}\n{self.z}""")  # date,totalMoney, n
            cache.close()
        with open('Config/checkWindow.ini', 'w+') as config:
            config.write(str(False))
            config.close()

        self.checkExcel()
        self.saveExit()
        exit()

    # report GET
    def saveExcel(self):
        try:

            DIR_PATH = str(datetime.now().strftime("%B")) + f"_{datetime.now().year}/"
            SAVE_PATH = '//Zerozed-pc/shared/DB/Sales_Report/'+DIR_PATH
            BASE_FILE = 'data/Report Data' + '/daily_report_example.xlsx'
            EXCEL_FILE = SAVE_PATH + str(date.today()) + '.xlsx'
            base = []

            # precaution if DIR_PATH not exist
            if not os.path.exists(os.path.abspath(SAVE_PATH)):
                try:
                    print(SAVE_PATH + " CREATED")
                    os.makedirs(os.path.dirname(SAVE_PATH))
                except OSError as exc:  # Guard against race condition
                    if exc.errno != errno.EEXIST:
                        raise

            # Precaution if BASE_FILE not found
            if not os.path.exists(os.path.abspath(BASE_FILE)):
                try:
                    wb = pyxl.Workbook()
                    ws = wb.active
                    ws.sheet_view.showGridLines = True
                    ws.title = "MAIN"
                    ws.column_dimensions['A'].width = 20
                    ws.column_dimensions['B'].width = 30
                    ws.column_dimensions['C'].width = 10
                    ws.column_dimensions['D'].width = 10
                    ws.column_dimensions['E'].width = 10
                    ws['D1'].alignment = Alignment(wrap_text=True)
                    ws['E1'].alignment = Alignment(wrap_text=True)
                    ws['A1'].value = 'Code'
                    ws['B1'].value = 'Name'
                    ws['C1'].value = 'Qty Sold'
                    ws['D1'].value = """Total Sales\nAmount\n(Gross RM)"""
                    ws['E1'].value = """Average\nUnit Price\nRM"""
                    wb.save(filename=BASE_FILE)
                    wb.close()
                except OSError as exc:  # Guard against race condition
                    if exc.errno != errno.EEXIST:
                        raise

            # OPEN AND COPY FILE
            if not os.path.exists(os.path.abspath(EXCEL_FILE)):
                try:
                    shutil.copyfile(os.path.abspath(BASE_FILE), os.path.abspath(EXCEL_FILE))
                except OSError as exc:  # Guard against race condition
                    if exc.errno != errno.EEXIST:
                        raise

            self.checkExcel()  # check if saveExit existed or not

            for child in self.buyScreen.get_children():
                base.append([self.buyScreen.item(child)["values"][i] for i in range(6)])

            # print(base)   #debug
            '''EXCEL FOR REPORT DATA A.Code B.Name C.Qty D.Total E.Avg'''
            theFile = pyxl.load_workbook(EXCEL_FILE)
            allSheetNames = theFile.sheetnames
            qty_adj = 'C'
            total_adj = 'D'
            avg_adj = 'E'

            for sheet in allSheetNames:
                # print("Current sheet name is {}".format(sheet))   #debug find sheet
                currentSheet = theFile[sheet]

            def find_cell(value="", show="row"):
                for row in range(1, currentSheet.max_row + 1):
                    for column in "ABCDE":  # Here you can add or reduce the columns
                        cell_name = "{}{}".format(column, row)

                        if currentSheet[cell_name].value == value:
                            # print("{1} cell is located on {0}" .format(cell_name, currentSheet[cell_name].value))
                            # print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))   #debug
                            # print(column, row)
                            if show.lower() == "row":
                                return row
                            elif show.lower() == "column":
                                return column
                            elif show.lower() == "cell":
                                return cell_name
                            else:
                                return 'Only row, column, cell'

            def add_cell(value=['x', 'x', 1, 1.1, 1.1], column="ABCDE"):
                '''value Must depends on columns like 5 value for "ABCDE" '''

                wb = xlrd.open_workbook(EXCEL_FILE)
                sheet = wb.sheet_by_index(0)

                check_row = sheet.nrows
                i = 0
                for col in column:  # Here you can add or reduce the columns
                    cell_name = "{}{}".format(col, (check_row + 1))
                    currentSheet[cell_name].value = value[i]
                    i += 1

                # print("Successfully add new item")  #debug
                theFile.save(EXCEL_FILE)
                theFile.close()

            k = 0
            for data in base:
                row_adj = str(find_cell(value=str(data[k]), show='row'))
                if row_adj == 'None':
                    add_cell(value=[str(data[0]), data[1], data[2], float(data[4]), float(data[4]) / float(data[2])])
                else:
                    currentSheet[qty_adj + row_adj].value = int(currentSheet[qty_adj + row_adj].value) + int(data[2])
                    currentSheet[total_adj + row_adj].value = float(
                        float(currentSheet[total_adj + row_adj].value) + (float(data[4]) - float(data[5])))
                    currentSheet[avg_adj + row_adj].value = (float(currentSheet[total_adj + row_adj].value) + float(
                        data[4]) - float(data[5])) / \
                                                            (int(currentSheet[qty_adj + row_adj].value) + int(data[2]))

                    theFile.save(EXCEL_FILE)
                    theFile.close()


        except Exception as e:
            messagebox.showerror("ERROR", f"{e} . PLS CONTACT PROVIDER TO FIX IT")

    def saveExit(self):
        try:
            DIR_PATH = str(datetime.now().strftime("%B")) + f"_{datetime.now().year}/"
            SAVE_PATH = '//Zerozed-pc/shared/DB/Sales_Report/' + DIR_PATH
            BASE_FILE = 'data/Report Data' + '/daily_report_example.xlsx'
            EXCEL_FILE = SAVE_PATH + str(date.today()) + '.xlsx'

            # precaution if DiR_PATH not exist
            if not os.path.exists(os.path.dirname(SAVE_PATH)):
                try:
                    os.makedirs(os.path.dirname(SAVE_PATH))
                except OSError as exc:  # Guard against race condition
                    if exc.errno != errno.EEXIST:
                        raise

            # Precaution if BASE_FILE not found
            if not os.path.exists(os.path.abspath(BASE_FILE)):
                try:
                    wb = pyxl.Workbook()
                    ws = wb.active
                    ws.sheet_view.showGridLines = True
                    ws.title = "MAIN"
                    ws.column_dimensions['A'].width = 20
                    ws.column_dimensions['B'].width = 30
                    ws.column_dimensions['C'].width = 10
                    ws.column_dimensions['D'].width = 10
                    ws.column_dimensions['E'].width = 10
                    ws['D1'].alignment = Alignment(wrap_text=True)
                    ws['E1'].alignment = Alignment(wrap_text=True)
                    ws['A1'].value = 'Code'
                    ws['B1'].value = 'Name'
                    ws['C1'].value = 'Qty Sold'
                    ws['D1'].value = """Total Sales\nAmount\n(Gross RM)"""
                    ws['E1'].value = """Average\nUnit Price\nRM"""
                    wb.save(filename=BASE_FILE)
                    wb.close()
                except OSError as exc:  # Guard against race condition
                    if exc.errno != errno.EEXIST:
                        raise

            # OPEN AND COPY FILE
            if not os.path.exists(os.path.abspath(EXCEL_FILE)):
                try:
                    shutil.copyfile(os.path.abspath(BASE_FILE), os.path.abspath(EXCEL_FILE))
                except OSError as exc:  # Guard against race condition
                    if exc.errno != errno.EEXIST:
                        raise

            # openpyxl
            theFile = pyxl.load_workbook(EXCEL_FILE)
            allSheetNames = theFile.sheetnames
            for sheet in allSheetNames:
                # print("Current sheet name is {}".format(sheet))   #debug find sheet
                currentSheet = theFile[sheet]

            # xlrd
            wb = xlrd.open_workbook(EXCEL_FILE)
            sheetwb = wb.sheet_by_index(0)

            target_row = sheetwb.nrows + 4

            raw_data = [value for value in currentSheet.iter_rows(min_row=1,
                                                                  max_row=sheetwb.nrows,
                                                                  min_col=1,
                                                                  max_col=5,
                                                                  values_only=True)]
            raw_data.pop(0)
            # pprint(raw_data)    #debug

            qty = price = 0
            for data in raw_data:
                # pprint(data)  #debug
                qty += data[2]
                price += data[3]
            # print(qty, price)   #debug

            currentSheet['A' + str(target_row)].alignment = Alignment(wrap_text=True)
            currentSheet['A' + str(target_row + 1)].alignment = Alignment(wrap_text=True)
            currentSheet['A' + str(target_row)].value = "Total Product\nSold Today"
            currentSheet['A' + str(target_row + 1)].value = "Total Cash\nToday(RM)"
            currentSheet['A' + str(target_row + 3)].value = "Date"
            currentSheet['B' + str(target_row)].value = qty
            currentSheet['B' + str(target_row + 1)].value = float(price)
            currentSheet['B' + str(target_row + 3)].value = date.today()
            theFile.save(EXCEL_FILE)
            theFile.close()

        except Exception as e:
            messagebox.showerror("ERROR", f"{e} . PLS CONTACT PROVIDER TO FIX IT")

    def checkExcel(self):

        DIR_PATH = str(datetime.now().strftime("%B")) + f"_{datetime.now().year}/"
        SAVE_PATH = '//Zerozed-pc/shared/DB/Sales_Report/' + DIR_PATH
        BASE_FILE = 'data/Report Data' + '/daily_report_example.xlsx'
        EXCEL_FILE = SAVE_PATH + str(date.today()) + '.xlsx'

        if os.path.exists(os.path.abspath(EXCEL_FILE)):
            theFile = pyxl.load_workbook(EXCEL_FILE)
            allSheetNames = theFile.sheetnames

            for sheet in allSheetNames:
                # print("Current sheet name is {}".format(sheet))   #debug find sheet
                currentSheet = theFile[sheet]

                # xlrd
                wb = xlrd.open_workbook(EXCEL_FILE)
                sheetwb = wb.sheet_by_index(0)

                raw_data = [value for value in currentSheet.iter_rows(min_row=1,
                                                                      max_row=sheetwb.nrows,
                                                                      min_col=1,
                                                                      max_col=5,
                                                                      values_only=True)]

                footer = 7
                header = sheetwb.nrows - footer
                # pprint(raw_data)  #debug

                for data in raw_data:
                    if 'Date' in data:
                        currentSheet.delete_rows(header + 1, footer)
                        print('Footer found and deleted')
                        theFile.save(EXCEL_FILE)
                        theFile.close()

        else:
            messagebox.showwarning("No Data Found", f"Report for {date.today()} saved but no data inside")

    def printExcel(self):
        self.calendar_win = Toplevel(self.cashWin)
        Label(self.calendar_win, text='Choose date').pack(padx=10, pady=10)
        style = Style()
        style.configure('W.TButton', font=('Comic sans ms', 15, 'normal', 'italic'), foreground='black')
        style1 = 'W.TButton'

        today = date.today()

        mindate = date(year=2010, month=1, day=1)
        maxdate = today
        check_validate = StringVar()
        validate = Label(self.calendar_win, text='Report Exist', textvariable=check_validate,
                         font=('comic sans ms', 13, 'bold'), foreground='Green')

        cal = tkcalendar.Calendar(self.calendar_win, font=('comic sans ms', 15, 'bold'), selectmode='day',
                                  locale='en_US',
                                  mindate=mindate, maxdate=maxdate, disabledforeground='red',
                                  cursor="hand1", year=today.year, month=today.month, day=today.day)

        def print_report():
            DIR_PATH = str(datetime.now().strftime("%B")) + f"_{datetime.now().year}/"
            SAVE_PATH = '//Zerozed-pc/shared/DB/Sales_Report/' + DIR_PATH
            get_date = cal.selection_get()
            EXCEL_FILE = SAVE_PATH + str(get_date) + '.xlsx'

            if os.path.exists(os.path.abspath(EXCEL_FILE)):
                printfile = os.path.abspath(EXCEL_FILE)
                self.checkExcel()
                self.saveExit()
                try:
                    try:
                        with open('Config/printerConfig.ini', 'r') as f:
                            check_print = [i.strip("\n") for i in f.readlines()]
                            win32print.SetDefaultPrinter(str(check_print[3]))
                            f.close()
                    except Exception as e:
                        messagebox.showerror("Printer Error",
                                             f"Cant Print, Please contact service provider +60179593309\n{e}")

                    os.startfile(rf"{printfile}", 'print')
                    self.calendar_win.destroy()
                    messagebox.showinfo("Printing", f"Report on {get_date} Printed Successfully")
                except Exception as error:
                    messagebox.showerror("Printer error", error)
                    self.calendar_win.destroy()
            else:
                messagebox.showerror("Not Found", "Your Report NOT FOUND")
                self.calendar_win.focus_force()

        def is_validate(event):
            DIR_PATH = str(datetime.now().strftime("%B")) + f"_{datetime.now().year}/"
            SAVE_PATH = '//Zerozed-pc/shared/DB/Sales_Report/' + DIR_PATH
            get_date = cal.selection_get()
            EXCEL_FILE = SAVE_PATH + str(get_date) + '.xlsx'

            if os.path.exists(os.path.abspath(EXCEL_FILE)):
                validate.config(foreground='green')
                check_validate.set(f"Report {get_date} Exist")
            else:
                validate.config(foreground='red')
                check_validate.set(f"Report {get_date} Not Exist")

        self.calendar_win.bind("<Button-1>", is_validate)
        cal.pack(fill="both", expand=True)
        Button(self.calendar_win, text="Print", style='W.TButton', command=print_report).pack()
        validate.pack()


# TODO add order window and function
# TODO will use and research for DELIVERY_FLOW
# TODO use FLOW_CACHE
# TODO try to think how to setup member data
# TODO small bug for code1

# CashierWin(Tk(), ID='RZ0000E001')  # debug for one cashier_win.py
